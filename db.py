import openpyxl
import os
from datetime import datetime


class Excel:
    def __init__(self):
        self.header = ['Ссылка', 'Категория курса', 'Статус курса', 'Дата создания темы', 'Название темы',
                                'Количество просмотров', 'Количество просмотров деленное на количестство дней на сайте',
                                'Цена', 'Взнос', 'Цена деленная на взнос', 'Количество Складчиков в основном списке',
                                'Количество просмотров деленное на количество складчиков', 'Хэштеги',
                                'Cреднее количество участников в день']
        self.file_name = 'data.xlsx'
        self.wb = None

    def create_file(self):
        self.wb = openpyxl.Workbook()
        ws = self.wb.active
        for column in range(1, len(self.header) + 1):
            ws.cell(row=1, column=column).value = self.header[column - 1]
        self.wb.save(self.file_name)

    def save(self):
        self.wb.save(self.file_name)

    def add_skladchina(self, data: dict):
        if not os.path.isfile(self.file_name):
            self.create_file()
        if not self.wb:
            self.wb = openpyxl.load_workbook(self.file_name)
        ws = self.wb.active
        max_row = ws.max_row
        ws.cell(row=max_row + 1, column=1).value = data['url']
        ws.cell(row=max_row + 1, column=2).value = data['rubric']
        ws.cell(row=max_row + 1, column=3).value = data['status']
        ws.cell(row=max_row + 1, column=4).value = data['date'].strftime("%d.%m.%Y")
        ws.cell(row=max_row + 1, column=5).value = data['name']
        views = float(data['views'].replace('.', ''))
        ws.cell(row=max_row + 1, column=6).value = views
        days_on_site = (datetime.now() - data['date']).days
        try:
            ws.cell(row=max_row + 1, column=7).value = views / days_on_site
        except ZeroDivisionError:
            ws.cell(row=max_row + 1, column=7).value = 'Деление на 0'
        price = float(data['price'].replace(' руб', ''))
        ws.cell(row=max_row + 1, column=8).value = price
        deposit = float(data['deposit'].replace(' руб', ''))
        ws.cell(row=max_row + 1, column=9).value = deposit
        ws.cell(row=max_row + 1, column=10).value = price / deposit
        main = float(data['main'].replace('.', ''))
        ws.cell(row=max_row + 1, column=11).value = main
        try:
            ws.cell(row=max_row + 1, column=12).value = views / main
        except ZeroDivisionError:
            ws.cell(row=max_row + 1, column=12).value = 'Деление на 0'
        ws.cell(row=max_row + 1, column=13).value = ', '.join(data['hash_tags'])
        if data['event']['count'] == 0:
            averape_people_per_day = 0
        else:
            try:
                averape_people_per_day = data['event']['count'] / (datetime.now() - data['event']['date']).days
            except ZeroDivisionError:
                averape_people_per_day = 0
        ws.cell(row=max_row + 1, column=14).value = averape_people_per_day
